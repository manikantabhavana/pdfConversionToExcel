# -*- coding: utf-8 -*-



"""
Created on Thu Oct  5 21:33:46 2023

@author: ManikantaBhavana@StepNext

Mail:manikanta@stepnext.com
"""


#main task is to extract the text from the scanned pdf
#import the required modules for pdf conversion

import re
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import string
import pytesseract
from pytesseract import Output
from PyPDF2 import PdfFileReader
import PyPDF2
from pdf2image import convert_from_path
import concurrent.futures
import cv2
import ocrmypdf
import datetime
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont,ImageEnhance
from openpyxl.drawing.image import Image as img
from tqdm import tqdm



# Function to preprocess an image
def preprocess_image(image):
    
    # Convert the image to grayscale
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # Apply thresholding to enhance text
    _, thresholded = cv2.threshold(gray, 2, 130, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    return Image.fromarray(thresholded)

def process_page(page_image,page_num):
    
    
    image=page_image
    sec_image=image.crop((10,38,1000,70))
    
    section=pytesseract.image_to_string(sec_image,config=Output.STRING)
    section=section.replace('Cartinn','Section')
    
    section=section.replace('Section No and Name','')
    section=section.replace(':','')
    section=section.strip()
    
    # Get the dimensions of the image
    image_width, image_height = image.size

    # Define the number of rows and columns in the matrix
    num_rows = 10
    num_columns = 3
    extracted_text=[]
    width=(image.width-120)/3-150
    width1=170
    
    for j in range(3):
       
        x1=35+j*(width+width1-10)
        
        y1=75
        
        x2=x1+width
        
        y2=image.height-50
        
        cell=image.crop((x1,y1,x2,y2))
        
        
        cell1=image.crop((x2-5,y1,x2+width1,y2))
        
        time=datetime.datetime.now()
        time=str(time)
        time=time[-7:-1]
        #cell.save('cells/'+time+'.jpg')
        custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ZYXWUTSRQPDMNLVKJOIHGFECBA0123456789'
        custom_config1 = r'--oem 3 --psm 6'
        # Perform OCR on the preprocessed image
        extracted_id = pytesseract.image_to_string(cell1, config=custom_config)
        words = re.findall(r'\S+',extracted_id)

# Find the length of the longest words
        max_length = 5

# Find all the longest words
        ids = [word for word in words if len(word) >= max_length]
        
        if(len(ids)==0):
            ids=['NID0000000']*10
        
        
        prefixed_ids = [f"{'Epic'} {id}" for id in ids]
        if(section==''):
            section='Additions'
        
        areas = [section]*len(prefixed_ids)
        
        prefixed_areas = [f"{'Area :'} {area}" for area in areas]
        
        cell=preprocess_image(cell)
        
        extracted_info = pytesseract.image_to_string(cell, config=custom_config1)
        extracted_info=extracted_info.replace('Nama :','Name :')
        extracted_info=extracted_info.replace('Narne =','Name :')
        extracted_info=extracted_info.replace('¥','V');
        extracted_info=extracted_info.replace('/','');
        
        extracted_info='\n'+extracted_info
        
        extracted_info=extracted_info.split('\nName')[1:11]
        
        prefixed_info = [f"{'Name'} {name}" for name in extracted_info]
        
        final_text = '\n'.join(f'{id}\n{info}\n{area}' for id, info,area in zip(prefixed_ids, prefixed_info,prefixed_areas))
        ids=','.join(id for id in ids)
        
        extracted_text.append(final_text)
    
        
    
    
    return extracted_text,section





def get_data(pdf):

    pdf_file = pdf
    images = convert_from_path(pdf_file)
    
    demo_image=images[0]
    
    
    # Create a list of pages to process
    pages_to_process = images[2:-1]
    
    # Initialize an empty list to store the extracted text from each page
    final_extracted_text = []
    areas=[]
    
    extracted_text_by_page = {}
    areas_by_page = {}
    
    # Use a ThreadPoolExecutor for concurrent processing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Process pages in parallel and store the Future objects
        futures = {executor.submit(process_page, page,page_number): page_number for page_number, page in enumerate(pages_to_process, start=1)}
    
        # Retrieve the results as they become available and store in the dictionary
        for future in concurrent.futures.as_completed(futures):
            page_number = futures[future]
            extracted_text,section = future.result()
            
            areas_by_page[page_number]=section
            
            extracted_text_by_page[page_number] = extracted_text
    
    # Concatenate the extracted text in page number order
    sorted_extracted_text = [extracted_text_by_page[page_number] for page_number in range(1, len(pages_to_process) + 1)]
    
    sorted_areas = [areas_by_page[page_number] for page_number in range(1, len(pages_to_process) + 1)]
    
    
    cells='' 
    
    final_extracted_text1=[col for sublist in sorted_extracted_text for col in sublist ]  
    
    for page in final_extracted_text1:
      
    # Remove any empty records or records with only whitespace
      
       cells=cells+page
       
    filtered_areas=set(sorted_areas)
    file_path = "output.txt"

# Open the file in write mode ('w')
    with open(file_path, 'w') as file:
    # Write the string to the file
        file.write(cells)
    
    

    
    return cells,sorted_areas,demo_image,filtered_areas    
    





def get_info_df(data):
    li=data.split('Epic')
    li1=li[1::]
    
    df=pd.DataFrame(li1)
    
    df[0]=df[0].str.replace('\n',' ')
    df[0]=df[0].str.replace('*',':')
    df[0]=df[0].str.replace('$',':')
    df[0]=df[0].str.replace('+',':')
    df[0]=df[0].str.replace('=',':')
    df[0]=df[0].str.replace('?',':')
    df[0]=df[0].str.replace('~',':')
    df[0]=df[0].str.replace('>>',':')
    df[0]=df[0].str.replace('>> ',':')
    df[0]=df[0].str.replace('»',':')
    df[0]=df[0].str.replace(':-',':')
    df[0]=df[0].str.replace('_','')
    df[0]=df[0].str.replace('Husbana:s','Husband’s')


    df[0]=df[0].str.replace('Name !','Name :')
    df[0]=df[0].str.replace('Nama :','Name :')
    df[0]=df[0].str.replace('Number !','Number :')
    
    df[0]=df[0].str.replace('Narne','Name')
    df[0]=df[0].str.replace('Gonder :','Gender :')
   
    df[0]=df[0].str.replace('Age !','Age :')
    df[0]=df[0].str.replace('Gender !','Gender !')
    df[0]=df[0].str.replace('Name ,','Name :')
    df[0]=df[0].str.replace('Number ,','Number :')
    df[0]=df[0].str.replace('Age ,','Age :')
    df[0]=df[0].str.replace('Gender ,','Gender !')
    df[0]=df[0].str.replace('Ags','Age :')
    
    df[0]=df[0].str.replace(";",':')
    df[0]=df[0].str.replace(">",':')
    df[0]=df[0].str.replace("<",':')
    df[0]=df[0].str.replace("‘",'')
    df[0]=df[0].str.replace("`",'')
    df[0]=df[0].str.replace(".",'')
    df[0]=df[0].str.replace('|','')
    df[0]=df[0].str.replace('¢',':')
    df[0]=df[0].str.replace('§','5')
    df[0]=df[0].str.replace('©','')
    df[0]=df[0].str.replace(',','')
    df[0]=df[0].str.replace("'",'')
    
    
    df[0]=df[0].str.replace('Number !','Number :')
    df[0]=df[0].str.replace('Age !','Age :')
    df[0]=df[0].str.replace('Gender !','Gender !')
    df[0]=df[0].str.replace('’',':')
    df[0]=df[0].str.replace('!','i')
    
    df[0]=df[0].str.replace('Husband’s Name','Fathers Name')
    df[0]=df[0].str.replace('Husband’s Nama','Fathers Name')
    df[0]=df[0].str.replace('Husband’s Nama:','Fathers Name:')
    df[0]=df[0].str.replace('Husband:s Name','Fathers Name')
    df[0]=df[0].str.replace('Husban:s Name','Fathers Name')
    df[0]=df[0].str.replace('Husbands Name','Fathers Name')
    df[0]=df[0].str.replace('Husbands Nama:','Fathers Name:')
    df[0]=df[0].str.replace('Father’s Nama','Fathers Name') 
    
    df[0]=df[0].str.replace('Mother’s Name','Fathers Name')
    df[0]=df[0].str.replace('Mother’s Nama','Fathers Name')
    df[0]=df[0].str.replace('Mothers Name','Fathers Name')
    df[0]=df[0].str.replace('Mothers Nama','Fathers Name')
    df[0]=df[0].str.replace('Wife’s Name','Fathers Name')
    df[0]=df[0].str.replace('Wifes Name','Fathers Name')

    df[0]=df[0].str.replace('Guru’s Name','Fathers Name')
    df[0]=df[0].str.replace('Gurus Name','Fathers Name')
    df[0]=df[0].str.replace('Gaurdian Name','Fathers Name')
    df[0]=df[0].str.replace('Others','Fathers Name')
    df[0]=df[0].str.replace('Other’s','Fathers Name')
    df[0]=df[0].str.replace('Fathers Nama','Fathers Name')
    df[0]=df[0].str.replace('Fathers Namo','Fathers Name')
    df[0]=df[0].str.replace('Gander','Gender')
    df[0]=df[0].str.replace('Gondor','Gender')
    df[0]=df[0].str.replace('Femata','Female')
    df[0]=df[0].str.replace('Femate','Female')
    df[0]=df[0].str.replace('Famale','Female')
    df[0]=df[0].str.replace('Femaie','Female')
    df[0]=df[0].str.replace('Fomale','Female')
    df[0]=df[0].str.replace('Famals','Female')
    df[0]=df[0].str.replace('Femaia','Female')
    df[0]=df[0].str.replace('Femals','Female')
    df[0]=df[0].str.replace('Femaie','Female')
    df[0]=df[0].str.replace('Maie','Male')
    df[0]=df[0].str.replace('Malo','Male')
    df[0]=df[0].str.replace('Mala','Male')
    df[0]=df[0].str.replace('Maio','Male')
    df[0]=df[0].str.replace('Mais','Male')
    df[0]=df[0].str.replace('Mate','Male')
    df[0]=df[0].str.replace('Maia','Male')
    df[0]=df[0].str.replace('Fernale','Female')
    df[0]=df[0].str.replace('Femala','Female')
    
    
  
    
    df[0]=df[0].str.replace('Housa','House')
    df[0]=df[0].str.replace('Hause','House')
    df[0]=df[0].str.replace('Numbgs','Number')
    df[0]=df[0].str.replace('Ags :','Age :')
    
    df[0]=df[0].str.replace('Legal','Fathers Name')
    df[0]=df[0].str.replace('Gaurdian','Fathers Name')
    df[0]=df[0].str.replace('gaurdian','Fathers Name')
    df[0]=df[0].str.replace('Photo is','')
    df[0]=df[0].str.replace('Available','')
    df[0]=df[0].str.replace('Availabl','')
    df[0]=df[0].str.replace('Availab','')
    df[0]=df[0].str.replace('Availabli','')
    df[0]=df[0].str.replace('::',':')
    df[0]=df[0].str.replace(': :',':')
    df[0]=df[0].str.replace(']','')
    df[0]=df[0].str.replace('[','')
    df[0]=df[0].str.replace('(','')
    df[0]=df[0].str.replace(')','')
    df[0]=df[0].str.replace('}','')
    df[0]=df[0].str.replace('{',':')
    df[0]=df[0].str.replace("\\\\",'/')
    
    df[0]=df[0].str.strip()
    df[0] = df[0].apply(lambda x: x.encode('utf-8', 'ignore').decode('utf-8'))
    df=pd.DataFrame(df)
    df.columns=['info']
    return df



def get_page_area_sno(df,areas):
    
    df['page_no'] = (df.index // 30) + 1
    
    
   
    
    
    
    
    df=df[df['info'].str.len()>10]
    page_counts = df.groupby('page_no').size()
    start_num=1
    s_no=[]
    
    for page_count in page_counts:
    # Take user input for the total number of elements (x)
        x=page_count
        
        # Ensure the matrix has exactly 30 elements
        if x > 30:
            x = 30
        
        # Create an array with sequential numbers up to x
        elements = np.arange(start_num, start_num+ x)
        
        # Pad the array with empty elements to reach 30 elements
        elements = np.append(elements, [''] * (30 - x))
        
        # Reshape the array into 10 rows and 3 columns
        matrix = elements.reshape(10, 3)
        numbers_column_wise = matrix.flatten('F')
        s=[int(i) for i in numbers_column_wise if str(i) >='1']
        sorted_s=sorted(s)
        start_num=sorted_s[-1]+1
        s_no=s_no+s
    
    
    df['S_NO']=s_no
    df = df.sort_values(by='S_NO', ascending=True)
    return df


def replace_dollar(text):
    
    # Split the text around the first '$'
    parts = re.split(r'(\$)', text, maxsplit=1)
    
    if len(parts) == 3:
        # Check if the split resulted in three parts
        prefix, _, suffix = parts
        
        # Check if the prefix or suffix contains alphabets
        if any(c.isalpha() for c in prefix) or any(c.isalpha() for c in suffix):
            
            return f"{prefix}S{suffix}"
        
        else:
            return f"{prefix}5{suffix}"
    else:
        # If there is no '$' or only one '$', return the text as is
        return text

# Apply the function to the 'info' column using a lambda function

def get_all_columns(df):

    # Define patterns to capture 'Name', 'Father Name', 'Age', and 'Gender'
    epic_pattern = r'([A-Z0-9]{1,4}[A-Z0-9]{6,10}\d)'
    name_pattern = r'Name\s*(?:\:)?\s*([\w\s\#/,\.\s\-&:;]+)Fathers'
    father_name_pattern = r'\s*Fathers Name\s*(?:\:)?\s*([\w\s\#/,\.\s\-&:;]+)House'
    #house_pattern = r'\s*House Number\s*(?:\:)?\s*([A-Za-z0-9#/,\.\s-]+)Age'
    #house_pattern = r'\s*House Number\s*(?:\:)?\s*([A-Za-z0-9#/,\.\s\-&:;_]+)Age'
    house_pattern = r'\s*House Number\s*(?:\:)?\s*(.+)Age'
    
    age_pattern = r'\s*Age\s*(?:\:)?\s*(\d+)'
    gender_pattern = r'\s*Gender\s*(?:\:)?\s*(\w+)'
    area_pattern = r'\s*Area\s*(?:\:)?\s*(.+)'
    
    # Use regular expressions to extract 'Name', 'Father Name', 'Age', and 'Gender'
    df['Epic'] = df['info'].apply(lambda x: re.search(epic_pattern, x).group(1) if re.search(epic_pattern, x) else None)
    df['Name'] = df['info'].apply(lambda x: re.search(name_pattern, x).group(1) if re.search(name_pattern, x) else None)
    df['Father Name'] = df['info'].apply(lambda x: re.search(father_name_pattern, x).group(1) if re.search(father_name_pattern, x) else None)
    df['House Number'] = df['info'].apply(lambda x: re.search(house_pattern, x).group(1) if re.search(house_pattern, x) else None)
    df['Age'] = df['info'].apply(lambda x: re.search(age_pattern, x).group(1) if re.search(age_pattern, x) else None)
    df['Gender'] = df['info'].apply(lambda x: re.search(gender_pattern, x).group(1) if re.search(gender_pattern, x) else None)
    df['Area'] = df['info'].apply(lambda x: re.search(area_pattern, x).group(1) if re.search(area_pattern, x) else None)
    df['Name']=df['Name'].str.upper()
    df['Father Name']=df['Father Name'].str.upper()
    df['House Number']=df['House Number'].str.strip()
    
    return df



def clean_age(age):
    if type(age)=='str' or type(age)=='int':
    
        if(int(age)<10):
            age=str(age)+str(1)
        elif len(str(age))>=3:
            age=str(age)
            age=age[:2]
        
    return age

def clean_epic(epic):
    if epic is not None:
        if(len(epic)==11):
            first_three_letters = ''.join(filter(str.isalpha, epic))
            seven_numbers=epic[-7:]
            
            
            new_epic=first_three_letters+seven_numbers
            return  new_epic
        else:
            return epic
    else:
        return ''


def extract_common_surname(row):
    if(row['Name'] is not None):
        name_tokens = row['Name'].split()
        if(len(name_tokens)>1):
            surname=name_tokens[-1]
            
        else:
            surname=''
    else:
        surname=''
            
    return surname
    
    
    
def categorize_age(age):
    if age is not None:
        age=int(age)
        if age < 30:
            return 'Young'
        elif 30 <= age <= 60:
            return 'Middle Age'
        else:
            return 'Senior Citizen'
    else:
        return ''



def get_demographics(demo_image):
    image=preprocess_image(demo_image)
    
    
    
    def clean(text):
        text=text.replace('>',':')
        text=text.replace('<',':')
        text=text.replace('=',':')
        text=text.replace('+',':')
        text=text.replace('}',':')
        text=text.replace('*',':')
        text=text.replace('@',':')
        text=text.replace('~',':')
        text=text.replace('!',':')
        text=text.replace('Mandal ','Mandal :')
        text=text.replace('District ','District :')
        text=text.replace('Pin coda','Pin Code')
        text=text.replace('::',':')
        return text
    
    assembly_cell=image.crop((83,200,1200,250)) 
    assembly=pytesseract.image_to_string(assembly_cell,config=r'--oem 3 --psm 6')
    assembly=clean(assembly)
    assembly=assembly.replace('\n',' ')
    assembly=assembly.replace('Constituency','Constituency :')
    assembly=assembly.replace(': :',':')
    
    assembly=assembly.split(':')
    if(len(assembly)>1):
        assembly=assembly[1]
    else:
        assembly=''
        
        
        
    booth_cell=image.crop((1380,200,1600,250))
    
    booth=pytesseract.image_to_string(booth_cell,config=r'--oem 3 --psm 6')
    
    booth=clean(booth)
    booth=booth.replace('Part No.','')
    booth=booth.replace(':','')
    booth=booth.replace('\n','')  
    booth=booth.replace(' ','')
    booth=booth.strip()
    
    
    mp_cell=image.crop((35,260,1600,335))
    mp=pytesseract.image_to_string(mp_cell,config=Output.STRING)
    mp=clean(mp)
    mp=mp.replace('\n','')
    mp=mp.replace('located','located :')
    mp=mp.replace(': :',':')
    mp=mp.split(':')
    if(len(mp)>1):
        mp=mp[1]
        mp=mp.strip()
    else:
        mp=''
        
        
    location_cell=image.crop((515,750,1500,1100))
    location=pytesseract.image_to_string(location_cell,config=r'--oem 3 --psm 6')
    location=clean(location)
    
    district=re.search(r'District\s*(?:\:)\s*(.+)',location,re.IGNORECASE)
    if district:
        district=district.group(1)
    else:
        district=''
    village=re.search(r'Village\s*(?:\:)\s*(.+)',location,re.IGNORECASE)
    if village:
        village=village.group(1)
    else:
        village=''
        
    mandal=re.search(r'Mandal\s*(?:\:)\s*(.+)',location,re.IGNORECASE)
    if mandal:
        mandal=mandal.group(1)
    else:
        mandal=''
        
    pincode=re.search(r'Pin\s*Code\s*(.+)',location,re.IGNORECASE)
    
    if pincode:
        pincode=pincode.group(1)
    else:
        pincode=''
    if(len(pincode)>6):
        pincode=pincode[-6:]
        
        
    polling_station_cell= image.crop((33,1500,855,1700))
    
    polling_station=pytesseract.image_to_string(polling_station_cell,config=r'--oem 3 --psm 6')
    
    polling_station=polling_station.split(':')
    if(len(polling_station)>1):
        polling_station=polling_station[1]
        polling_station=polling_station.strip()
    else:
        polling_station=''
    
    
    
    
    
    polling_station=polling_station.strip()
    
    
    demo_data=[assembly,mandal,village,pincode,polling_station,booth,mp,district]
    
   


    
    
  
    
    
    return demo_data
    
    
    


def get_demographics_df(df,demo_data):

    demographics=df[['S_NO','Epic','Name','Father Name','Surname','House Number','Age','Gender','Area']]
    
    
    assembly=demo_data[0]
    demographics['Constituency']=assembly
    mandal=demo_data[1]
    demographics['Mandal']=mandal
    village=demo_data[2]
    demographics['Village']=village
    pincode=demo_data[3]
    demographics['Pincode']=pincode
    polling_station=demo_data[4]
    demographics['Polling_Station']=polling_station
    booth=demo_data[5]
    booth=booth.replace('|','1')
    demographics['Booth']=booth
    mp=demo_data[6]
    demographics['Mp']=mp
    district=demo_data[7]
    demographics['District']=district
    
    return demographics


def clean_value(cell_value):
    if isinstance(cell_value, bytes):
        # If the value is bytes, clean it
        pattern = rb'[\x00-\x1F\x7F-\x9F]'
        cleaned_bytes = re.sub(pattern, b'', cell_value)
        return cleaned_bytes
    elif isinstance(cell_value, str):
        # If the value is a string, clean it
        pattern = r'[\x00-\x1F\x7F-\x9F]'
        cleaned_string = re.sub(pattern, '', cell_value)
        return cleaned_string
    else:
        # Return other data types as is
        return cell_value



def get_family(df):

    grp_house=df.groupby(['House Number'])
    srt=[]
    for i,grp in grp_house:
        sort_grp=grp.sort_values(by='Age',ascending=False)
        srt.append(sort_grp)
        
    df1=pd.DataFrame()
    srt.sort(key=len,reverse=True)
    for s in srt:
        if(len(s)>1):
            s['Head']=range(len(s))
       
        df1=pd.concat([df1,s],ignore_index=True)
    
    
    row_no=[]
    df1['S_NO']=range(1,len(df1)+1)
    n=1
    for h in df1['Head']:
        n=n+1
        if h==0:
            
            row_no.append(n)

    

    df1['Head']=df1['Head'].fillna('Individual')
            
    return row_no,df1




def get_demoTable(demo_data,df,filtered_areas):
    
    areas=filtered_areas
    areass=''
    for area in areas:
        areass=areass +' ,'+area
    
    male_count=len(df[df['Gender']=='MALE'])
    
    female_count=len(df[df['Gender']=='FEMALE'])
    
    others_count=len(df)-(male_count+female_count)
    
    assembly,mandal,village,pincode,polling_station,booth,mp,district=demo_data
    
    df1=pd.DataFrame({'Constituency':assembly,'MP_Constituency':mp,'Booth NO':booth,'Polling Station':polling_station},index=['1'])
    
    df2=pd.DataFrame({'Voters':len(df),'Male Voters':male_count,'Female Voters':female_count,'Others':others_count},index=['1'])
    
    df3=pd.DataFrame({'Village':village,'Mandal':mandal,'Pincode':pincode,'Areas':areass},index=[1])
    
    return [df1,df2,df3]




def analytics(c_demo,voters_df):
    df=c_demo
    
    
    
    counts=[voters_df['Male Voters'][0],voters_df['Female Voters'][0],voters_df['Others'][0]]
    
    
    plt.figure(figsize=(6, 6))
    plt.pie(counts, labels=['Male','Female','Third Gender'], autopct='%1.1f%%', startangle=140)
    plt.title('Gender Distribution')

# Save the pie chart as an image (e.g., PNG)
    plt.savefig('gender_pie_chart.png')
    
    age_group_counts = df['AgeGroup'].value_counts()
    plt.figure(figsize=(6, 6))
    plt.pie(age_group_counts, labels=age_group_counts.index, autopct='%1.1f%%', startangle=140)
    plt.title('Age Group Distribution')
    plt.savefig('age_group_pie_chart.png')
    # Pivot the DataFrame to count occurrences of each combination
    pivot_df = df.pivot_table(index='AgeGroup', columns='Gender', aggfunc='size', fill_value=0)
    age_groups = pivot_df.index
    genders = pivot_df.columns
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c']

    # Set the bar width and create an index array for the x-axis
    bar_width = 0.2
    x = np.arange(len(age_groups))

    # Create subplots for each gender
    fig, ax = plt.subplots(figsize=(15, 10))

    for i, gender in enumerate(genders):
        ax.bar(x + i * bar_width, pivot_df[gender], width=bar_width, label=gender, color=colors[i])


    # Set x-axis labels and title
    ax.set_xlabel('Age Groups')
    ax.set_ylabel('Count')
    ax.set_title('Age Group Distribution by Gender')

    # Set x-axis ticks and labels
    ax.set_xticks(x + (len(genders) - 1) * bar_width / 2)
    ax.set_xticklabels(age_groups)

    # Add a legend
    ax.legend()
    plt.savefig('age.png')
    # Create a workbook and add a worksheet
    



def get_excel_file(raw,main_data,cl_demographics,c_family,list_details_df,row_no):
    
    constituency_df,voters_df,area_df=list_details_df
    
    constituency_df=constituency_df.applymap(clean_value)
    
    voters_df=voters_df.applymap(clean_value)
    
    area_df=area_df.applymap(clean_value)
    #constituency_df.to_excel('consti.xlsx')
    constituency_df['Booth NO']=constituency_df['Booth NO'].str.replace('|','1')

    file_name=constituency_df['Constituency'][0]+'-'+constituency_df['Booth NO'][0]+'.xlsx'
    
    excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')
    
    raw.to_excel(excel_writer, sheet_name='RAW_DATA', index=False, header=True)

    main_data.to_excel(excel_writer, sheet_name='FORMAT_DATA', index=False, header=True)

    cl_demographics.to_excel(excel_writer, sheet_name='DEMOGRAPHICAL_DATA', index=False, header=True)

    #c_family.to_excel(excel_writer,sheet_name='FAMILY_WISE_DATA',index=False,header=True)
    
    start_row=1
    
    constituency_df.to_excel(excel_writer,sheet_name='DETAILS',index=False,header=True,startrow=start_row)
    
    start_row=len(constituency_df)+4+start_row
    
    voters_df.to_excel(excel_writer,sheet_name='DETAILS',index=False,header=True,startrow=start_row)
    
    start_row=start_row+len(constituency_df)+2+len(voters_df)
    
    area_df.to_excel(excel_writer,sheet_name='DETAILS',index=False,header=True,startrow=start_row)
    
    # Get the workbook and the worksheet
    workbook = excel_writer.book
    
    worksheet = excel_writer.sheets['FORMAT_DATA']
    
    worksheet1=excel_writer.sheets['DEMOGRAPHICAL_DATA']
    
    #worksheet2=excel_writer.sheets['FAMILY_WISE_DATA']
    
    # Create a fill object with the desired color (e.g., yellow fill)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply the fill color to each cell in the header row
    for cell in worksheet[1]:
        
        cell.fill = yellow_fill
        
    for cell in worksheet1[1]:
        
        cell.fill = yellow_fill
        
    """for j in row_no:
        
        for cell in worksheet2[j]:
            
            cell.fill = yellow_fill"""
     
            
    
    excel_writer.save()
    # Write the Age data to the worksheet

    
    # Add the Age Group chart image to the worksheet
    

    # Save the Excel file
    
    # Save the Excel file
    

    """existing_workbook = load_workbook(file_name)

# Create a new sheet in the existing workbook

    ws = existing_workbook.create_sheet('ANALYTICS')
    age_group_img = img('age_group_pie_chart.png')
    ws.add_image(age_group_img, 'C3')

    # Generate the gender pie chart (assuming you already have it) and save it as 'gender_pie_chart.png'

    # Add the Gender chart image to the worksheet with a gap
    gender_img = img('gender_pie_chart.png')
    ws.add_image(gender_img, 'J3')  # Adjust the starting cell position


    gender_img = img('age.png')
    ws.add_image(gender_img, 'C25')  


    existing_workbook.save(file_name)"""





def convert(pdf):
    print('\n')
    
    print('************************************')
    print('\n')
    print('Pdf To Excel Converting....')
    
    print('Starts......')
    print('\n')
    print('This process will be completed in just a few seconds......')

      
    data,areas,demo_image,filtered_areas=get_data(pdf)
    
    demo_data=get_demographics(demo_image)
    
    df=get_info_df(data)
    
    df=get_page_area_sno(df,areas)   
        
    df['info'] = df['info'].apply(lambda x: replace_dollar(x)) 
     
    df=get_all_columns(df)
    
    df['Epic'] = df['Epic'].apply(lambda x: clean_epic(x))
    df['Age'] = df['Age'].apply(clean_age)
    
    df['Surname'] = df.apply(extract_common_surname, axis=1)
    
    
    
    demographics=get_demographics_df(df, demo_data)
    
    demographics['AgeGroup'] = demographics['Age'].apply(categorize_age)
    
    raw=df['info']
    
    main_data=df[['S_NO','Epic','Name','Father Name','Surname','House Number','Age','Gender','Area']]
    
    main_data=main_data.applymap(clean_value)
    
    row_no,family=get_family(main_data)
    
    # Apply the clean_bytes function to all cells in the DataFrame
    cl_demographics = demographics.applymap(clean_value)
    
    raw=raw.apply(clean_value)
    
    c_family=family.applymap(clean_value)
    
    main_data=main_data.applymap(clean_value)
    
    list_details_df=get_demoTable(demo_data, df, filtered_areas)
    
    #analytics(cl_demographics,list_details_df[1])
    
    get_excel_file(raw,main_data,cl_demographics,c_family,list_details_df,row_no)
    print('\n')
    print('\n')
    print('************************************')
    print('_____________100%___________________')
    print('Completd!!!')
    print('Booth PDF Converts Successfully !!!!!!')
    print('_____________________________________')
    print('Chech The Directory of input PDF For Output Excel File')



def SearchablePDF(input_pdf_file):



    print('Pdf Is Encrypted....???????????')
    print('************************************')
    print('PDF Decryption')
    print('************************************')

    with open(input_pdf_file, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfFileReader(pdf_file)

        # Create a new PDF file for the reversed pages
        pdf_writer = PyPDF2.PdfFileWriter()

        # Reverse and add pages to the new PDF
        for page_num in range(pdf_reader.getNumPages()):
            page = pdf_reader.getPage(page_num)
            pdf_writer.addPage(page)

        # Save the reversed PDF
        with open('decrypted.pdf', 'wb') as reversed_pdf:
            pdf_writer.write(reversed_pdf)



    print('Converts Scanned PDF Into A Searchable PDF')
    print('\n')

    print('************************************')
    print('\n')
    


    ocrmypdf.ocr('decrypted.pdf','ocr'+input_pdf_file)


def addAnalytics(excel_file):

    print('\n')
    print('___________________________________')
    print('processing........................')
    print('analysis starts.....................')

    
    
    

    
    main_data=pd.read_excel(excel_file,sheet_name='FORMAT_DATA')
    raw=pd.read_excel(excel_file,sheet_name='RAW_DATA')
    
    
    cl_demographics = pd.read_excel(excel_file,sheet_name='DEMOGRAPHICAL_DATA')


    
    #details= pd.read_excel(excel_file,sheet_name='DETAILS')
    
    
    
    main_data['Age']=main_data['Age'].apply(clean_age)
    cl_demographics['Age']=cl_demographics['Age'].apply(clean_age)

    row_no,family=get_family(cl_demographics)
    c_family=family.apply(clean_value)
    

    
    
    
    male_count=len(cl_demographics[cl_demographics['Gender']=='MALE'])
    female_count=len(cl_demographics[cl_demographics['Gender']=='FEMALE'])
    others=len(cl_demographics)-(male_count+female_count)
    data=pd.DataFrame({'Male Voters':male_count,'Female Voters':female_count,'Others':others},index=['1'])
    analytics(cl_demographics,data)
    excel_writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    
    raw.to_excel(excel_writer, sheet_name='RAW_DATA', index=False, header=True)

    main_data.to_excel(excel_writer, sheet_name='FORMAT_DATA', index=False, header=True)
    
    cl_demographics.to_excel(excel_writer,sheet_name='DEMOGRAPHICAL_DATA',index=False, header=True)
    
    c_family.to_excel(excel_writer,sheet_name='FAMILY_WISE_DATA',index=False,header=True)

    #details.to_excel(excel_writer,sheet_name='DETAILS',index=False,header=True)
    
    
    
    
    
    workbook = excel_writer.book
    
    worksheet = excel_writer.sheets['FORMAT_DATA']
    
    worksheet1=excel_writer.sheets['DEMOGRAPHICAL_DATA']
    
    worksheet2=excel_writer.sheets['FAMILY_WISE_DATA']
    
    #worksheet2=excel_writer.sheets['FAMILY_WISE_DATA']
    
    # Create a fill object with the desired color (e.g., yellow fill)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply the fill color to each cell in the header row
    for cell in worksheet[1]:
        
        cell.fill = yellow_fill
        
    for cell in worksheet1[1]:
        
        cell.fill = yellow_fill

    for j in row_no:
        for cell in worksheet2[j]:
            cell.fill=yellow_fill
    
    
    excel_writer.save()
    
    existing_workbook = load_workbook(excel_file)

# Create a new sheet in the existing workbook

    ws = existing_workbook.create_sheet('ANALYTICS')

    age_group_img = img('age_group_pie_chart.png')
    ws.add_image(age_group_img, 'C3')

    # Generate the gender pie chart (assuming you already have it) and save it as 'gender_pie_chart.png'

    # Add the Gender chart image to the worksheet with a gap
    gender_img = img('gender_pie_chart.png')
    ws.add_image(gender_img, 'J3')  # Adjust the starting cell position


    gender_img = img('age.png')
    ws.add_image(gender_img, 'C25')  

    excel_writer.book=existing_workbook
    data['Total Voters']=len(cl_demographics)
    data.to_excel(excel_writer,sheet_name='ANALYTICS',startrow=4,startcol=18,index=False)

    excel_writer.save()


    existing_workbook.save(excel_file)




    print("100'%' Completed....")
    print("Added All The Sheets To  Current Excel File ")
    print('☞☞☞☞☞☞'+excel_file)




def consolidate():
    list_files=os.listdir()
    consolidated_df=pd.DataFrame()
    family_df=pd.DataFrame()
    count_file=pd.DataFrame(columns=['Booth_no','count'])

    with tqdm(total=len(list_files),desc="Consolidating Files...") as pbar:
        for file in list_files:
            print('______________________________________')
            print('Files Consolidating..............')

            df=pd.read_excel(file,sheet_name='DEMOGRAPHICAL_DATA')
            df1=pd.read_excel(file,sheet_name='FAMILY_WISE_DATA')
            booth=df['Booth'][1]

            count=len(df)
            count_file=pd.concat([count_file,pd.DataFrame({'Booth_no':booth,'count':count},index=[0])],ignore_index=True)

            consolidated_df=pd.concat([consolidated_df,df],ignore_index=True)

            family_df=pd.concat([family_df,df1],ignore_index=True)

            pbar.update(1)

    print('Consolidation Complete...............')
    print('_____________________________________')
    consolidated_df.to_csv('test.csv')
    

    consolidated_df=consolidated_df.sort_values('Booth')

    consolidated_df['S_NO']=range(1,len(consolidated_df)+1)

    print('Write it into excel file .............')
    
    consolidated_df.to_excel('Consolidated.xlsx',index=False)

    family_df.to_excel('Family_Consolidated.xlsx',index=False)

    count_file.to_excel('count_file.xlsx',index=False)
    print('Consolidated File Is Saved In The Current Directory')
    print('File Name is Consolidated.xlsx')
    print('Family Head File Name is Family_Consolidated.xlsx')
    print('Count File Name is count_file.xlsx')






#SearchablePDF('Booth-75MMD.pdf')










